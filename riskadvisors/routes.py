from flask import url_for
from flask import request, redirect, session, render_template, jsonify
from sqlalchemy import Table, Column, Integer, String, MetaData, create_engine
from sqlalchemy.orm import mapper, create_session, clear_mappers
import uuid

import os
from riskadvisors import app,db

e=create_engine(app.config['SQLALCHEMY_DATABASE_URI']) 
db_session = create_session(bind=e, autocommit=False, autoflush=False)
    
@app.route('/', methods=['GET','POST'])
def index():
    try:
        if request.method == 'POST' and  'file_url' in request.form:
            file_url = request.form['file_url']
            file_url = file_url[0:-1]+'1'
            session['table_name']=request.form['table_name']
            return redirect(url_for('dropbox_handle', file_url=file_url))
            
        elif request.method == 'POST':
            session['table_name']=request.form['table_name']
            file = request.files['file']
            filename  = file.filename
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            session['filename']=filename
            return redirect(url_for('after_upload'))
        return render_template('index.html', table_name=str(uuid.uuid4()).replace('-',""))
    except:
        return "error occourred! return to home: <a href='"+url_for('index')+"'>Home<a/>"
    
@app.route('/dropbox_handle/')
def dropbox_handle():
    try:
        import wget
        file_url = request.args.get('file_url')
        filename = "f1.xlsx"
        f = wget.download(file_url,os.path.join(app.config['UPLOAD_FOLDER'],filename))
        session['filename']=f
        return redirect(url_for('after_upload'))
    except:
        return "Enter valid file url"+" return to home: <a href='"+url_for('index')+"'>Home<a/>"

class sheet(object):
    pass


@app.route('/after_upload/')
def after_upload():
        filename=session['filename']
        from openpyxl import load_workbook
        wb = load_workbook(filename=os.path.join(app.config['UPLOAD_FOLDER'],filename), read_only=True)
        #wb = load_workbook(filename='C://users/navdeep/Desktop/book.xlsx', read_only=True)
        
        ws = wb.active

        sheet_headers = []
        row_count=0
        for row in ws.rows:
            for cell in row:
                sheet_headers.append(cell.value)
            break
        for row in ws.rows:
            row_count+=1
        session['row_count'] = row_count;
        session['sheet_headers']=sheet_headers
        session['handle_size']=5000
        session['handler_count']=1

        return redirect(url_for('db_model'))
        
        
@app.route('/db_model', methods=['GET','POST'])
def db_model():
        if request.method == 'POST':
            sheet_headers = session['sheet_headers']
            tab=session['table_name']
            #e=create_engine(app.config['SQLALCHEMY_DATABASE_URI'])
            metadata = MetaData(bind=e)
            t = Table(tab, metadata, Column('id', Integer, primary_key=True),*(Column(header, String(8000)) for header in sheet_headers))
            metadata.create_all()
            
            return redirect(url_for('database_handler'))
            
        return  '''
            <!doctype html>
            <h1>Creating Database Schema for Table</h1>
            <form action="" method=post>
                <input id="autoclick" style='visibility:hidden;' type=submit value=Proceed>
            </form>
            <script>
            document.getElementById("autoclick").click();
            </script>
            '''
        

@app.route('/db_commit')
def db_commit():
    from openpyxl import load_workbook
    wb = load_workbook(filename=os.path.join(app.config['UPLOAD_FOLDER'],session['filename']), read_only=True)
    sheet_headers=session['sheet_headers']
    
    #wb = load_workbook(filename='C://users/navdeep/Desktop/book.xlsx', read_only=True)
    ws = wb.active

    tab=session['table_name']
    metadata = MetaData(bind=e)    
    t = Table(tab, metadata, Column('id', Integer, primary_key=True),*(Column(header, String(8000)) for header in sheet_headers))
    clear_mappers() 
    mapper(sheet, t)
      
    handler_count = session['handler_count']
    handle_size = session['handle_size']
    count=0
    handle_size_counter=0
    for r in ws.rows:
        count+=1
        if handler_count>count-1:
            continue
        else:
            handle_size_counter+=1
            #if count%100==0: print count,handle_size_counter,handle_size
            s = sheet()
            cou=0
            for c in r:
                setattr(s,sheet_headers[cou],c.value)
                cou+=1
                
            db_session.add(s)
            if handle_size_counter-1==handle_size:
                break
    session['handler_count']=handler_count+handle_size_counter
    db_session.commit()
    return redirect(url_for('database_handler'))

@app.route("/database_handler", methods=['GET','POST'])
def database_handler():
    

    if request.method == 'POST' and  session['handler_count']<session['row_count']:
        return redirect(url_for('db_commit'))
            
    elif request.method == 'POST' and session['handler_count']>=session['row_count']:
        return redirect(url_for('tableInfo'))
        


    if (int(session['handle_size'])*100)/int(session['row_count'])<15: msg="Looks like a big file, wait a minute"
    else: msg=''
    return  '''
            <!doctype html>
            <h1>Please Wait! Saving table to Database</h1>
            <p>'''+msg+'''</p>
            <h2>Completed '''+str((int(session['handler_count'])*100)/int(session['row_count']))+'''%</h2>
            <form action="" method=post>
                <input id="autoclick" style='visibility:hidden;' type=submit value=Proceed>
            </form>
            <script>
            document.getElementById("autoclick").click();
            </script>
            '''
@app.route("/tableInfo/")
def tableInfo():
    return  render_template('tableInfo.html')
@app.route("/queryPage")
def queryPage():
    return render_template('queryPage.html')

@app.route("/queryDb", methods=['GET','POST'])
def queryDb():
        metadata = MetaData(bind=e)    
        t = Table(session['table_name'], metadata, Column('id', Integer, primary_key=True),*(Column(header, String(8000)) for header in session['sheet_headers']))
        clear_mappers() 
        mapper(sheet, t)
        qu = request.get_json()
        print qu
        query_var_byUser = qu['col_value']
        query_column_byUser = qu['col_name']
        qargs = {query_column_byUser:query_var_byUser}
        acc = db_session.query(sheet).filter_by(**qargs).all()
        result_list=[]
        for item in acc:
            temp_dict={}
            for head in session['sheet_headers']:
                temp_dict[head]=getattr(item,head)
            result_list.append(temp_dict)

        print result_list
        return jsonify(result_list), 200